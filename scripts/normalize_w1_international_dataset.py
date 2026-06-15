#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""W1 S1B-Seed international dataset normalizer.

Reads the 4 sheets of the uploaded World Cup workbook and produces a single
unified, leakage-aware CSV plus coverage/quality summaries.

Discipline (W1_S0_SAFE_OUTPUT_AND_S1B_SEED_V1):
  * 90-minute score only for modeling (HGFT/AGFT or HG/AG); ET/penalties split out.
  * finish_type derived from ET/Pens columns, NOT from the dirty `Finished` label.
  * duplicate HGP header -> home_penalties / away_penalties.
  * team names mapped to canonical team_id via config/w1_team_aliases.json;
    unmapped names are a HARD ERROR (no silent fallback).
  * OU/AH absent in this source -> pipeline_mode = "1X2_ONLY",
    ou_market_available = ah_market_available = false, w1_full_pipeline_validated = false.

Raw .xlsx and the generated CSV are gitignored by default (see RESULT report).
"""
from __future__ import annotations

import argparse
import csv
import json
import re
import unicodedata
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = ROOT / "data/raw/international/WorldCup2026.xlsx"
ALIASES = ROOT / "config/w1_team_aliases.json"
OUT_CSV = ROOT / "data/processed/international/w1_international_dataset.csv"
OUT_SUMMARY = ROOT / "data/processed/international/w1_international_coverage.json"
CARDS_DIR = ROOT / "data/processed/match_cards/group_stage_round1"

SHEETS = ["WorldCup2026Qualifiers", "WorldCup2022", "WorldCup2018", "WorldCup2014"]
EDITION_HOST = {"WorldCup2014": "Brazil", "WorldCup2018": "Russia", "WorldCup2022": "Qatar"}
HOST_2026 = {"USA", "Mexico", "Canada"}

# Country-level synonym merges (applied AFTER accent stripping). Keep distinct
# countries separate (e.g. Congo vs DR Congo).
SYNONYMS = {
    "turkiye": "turkey",
    "czechia": "czech republic",
    "korea republic": "south korea",
    "republic of korea": "south korea",
    "korea dpr": "north korea",
    "dpr korea": "north korea",
    "united states": "usa",
    "united states of america": "usa",
    "cote d'ivoire": "ivory coast",
    "cote divoire": "ivory coast",
    "ivory coast": "ivory coast",
    "cape verde islands": "cape verde",
    "d.r. congo": "dr congo",
    "dr congo": "dr congo",
    "congo dr": "dr congo",
    "republic of ireland": "ireland",
    "ir iran": "iran",
    "iran": "iran",
    "china pr": "china",
    "bosnia and herzegovina": "bosnia & herzegovina",
}


def strip_accents(s: str) -> str:
    return "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))


def canon_name(raw: str) -> str:
    base = strip_accents(str(raw or "").strip())
    key = re.sub(r"\s+", " ", base).strip().lower()
    key = SYNONYMS.get(key, key)
    # Title-case canonical display, preserving '&'
    return " ".join(w if w in {"&"} else w.capitalize() for w in key.split())


def team_id_of(canonical: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", strip_accents(canonical).lower()).strip("_")


def w1_fixture_team_names() -> set[str]:
    names: set[str] = set()
    for p in sorted(CARDS_DIR.glob("*.json")):
        try:
            c = json.loads(p.read_text(encoding="utf-8"))
        except Exception:
            continue
        for side in ("home", "away"):
            nm = (c.get("teams", {}).get(side, {}) or {}).get("name")
            if nm:
                names.add(str(nm))
    return names


def load_workbook(xlsx: Path):
    import openpyxl
    return openpyxl.load_workbook(xlsx, read_only=True, data_only=True)


def sheet_rows(ws) -> list[dict[str, Any]]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(h) if h is not None else "" for h in rows[0]]
    # disambiguate duplicate headers (e.g. HGP, HGP -> HGP, HGP.1)
    seen: dict[str, int] = {}
    cols: list[str] = []
    for h in header:
        if h in seen:
            seen[h] += 1
            cols.append(f"{h}.{seen[h]}")
        else:
            seen[h] = 0
            cols.append(h)
    out = []
    for r in rows[1:]:
        if all(v is None for v in r):
            continue
        out.append({cols[i]: (r[i] if i < len(r) else None) for i in range(len(cols))})
    return out


def collect_all_names(wb) -> set[str]:
    names: set[str] = set()
    for s in SHEETS:
        if s not in wb.sheetnames:
            continue
        for row in sheet_rows(wb[s]):
            for k in ("Home", "Away"):
                if row.get(k):
                    names.add(str(row[k]).strip())
    return names


def build_aliases(xlsx: Path) -> dict[str, Any]:
    wb = load_workbook(xlsx)
    names = collect_all_names(wb) | w1_fixture_team_names()
    table: dict[str, Any] = {}
    for raw in sorted(names):
        canonical = canon_name(raw)
        tid = team_id_of(canonical)
        entry = table.setdefault(tid, {
            "team_id": tid, "canonical_name": canonical,
            "aliases": [], "fifa_code": None, "confederation": None,
            "host_auto_qualified_2026": tid in {"usa", "mexico", "canada"},
        })
        if raw not in entry["aliases"]:
            entry["aliases"].append(raw)
        entry["aliases"].sort()
    payload = {
        "schema_version": "W1_TEAM_ALIASES_V1",
        "note": "Auto-seeded from S1B seed workbook + W1 fixtures; canonical merges in normalizer SYNONYMS. Unmapped names are a hard error in the normalizer (no silent fallback).",
        "teams": dict(sorted(table.items())),
    }
    return payload


def load_alias_index() -> dict[str, str]:
    data = json.loads(ALIASES.read_text(encoding="utf-8"))
    idx: dict[str, str] = {}
    for tid, e in data.get("teams", {}).items():
        idx[e["canonical_name"]] = tid
        for a in e.get("aliases", []):
            idx[str(a)] = tid
        idx[tid] = tid
    return idx


def _f(v: Any) -> float | None:
    if v in (None, ""):
        return None
    try:
        x = float(v)
    except (TypeError, ValueError):
        return None
    return x


def _date(v: Any) -> str | None:
    if v is None or v == "":
        return None
    s = str(v)
    m = re.match(r"(\d{4}-\d{2}-\d{2})", s)
    return m.group(1) if m else s[:10]


def odds_1x2(row: dict[str, Any], finals: bool) -> tuple[float | None, float | None, float | None]:
    if finals:
        h = _f(row.get("H-Avg")) or _f(row.get("bet365-H")) or _f(row.get("Pinny-H")) or _f(row.get("H-Max"))
        d = _f(row.get("D-Avg")) or _f(row.get("bet365-D")) or _f(row.get("Pinny-D")) or _f(row.get("D-Max"))
        a = _f(row.get("A-Avg")) or _f(row.get("bet365-A")) or _f(row.get("Pinny-A")) or _f(row.get("A-Max"))
    else:
        h = _f(row.get("H_Avg")) or _f(row.get("H_Max"))
        d = _f(row.get("D_Avg")) or _f(row.get("D_Max"))
        a = _f(row.get("A_Avg")) or _f(row.get("A_Max"))
    return h, d, a


FIELDS = [
    "source_sheet", "competition", "season", "match_date", "phase", "stage",
    "neutral_site", "is_host_home", "is_host_away",
    "home_name_raw", "away_name_raw", "home_team_id", "away_team_id",
    "home_goals_90", "away_goals_90", "finish_type", "finished_label_raw", "dirty_finished_label",
    "home_goals_et", "away_goals_et", "home_penalties", "away_penalties",
    "odds_1x2_home", "odds_1x2_draw", "odds_1x2_away", "odds_1x2_available", "odds_source",
    "ou_main_line", "ou_over_odds", "ou_under_odds", "ou_market_available",
    "ah_main_line", "ah_market_available",
    "home_xg", "away_xg", "xg_available",
    "home_shots", "away_shots", "home_sot", "away_sot", "home_corners", "away_corners",
    "home_fouls", "away_fouls", "stats_available",
    "pipeline_mode", "w1_full_pipeline_validated", "prematch_available", "result_available",
]


def normalize(xlsx: Path) -> dict[str, Any]:
    wb = load_workbook(xlsx)
    idx = load_alias_index()
    rows_out: list[dict[str, Any]] = []
    unmapped: dict[str, int] = {}

    def mid(name: str) -> str | None:
        nm = str(name).strip()
        tid = idx.get(nm) or idx.get(canon_name(nm))
        if not tid:
            unmapped[nm] = unmapped.get(nm, 0) + 1
            return None
        return tid

    for sheet in SHEETS:
        if sheet not in wb.sheetnames:
            continue
        finals = sheet != "WorldCup2026Qualifiers"
        host = EDITION_HOST.get(sheet)
        season = int(re.sub(r"\D", "", sheet)) if re.search(r"\d", sheet) else None
        for row in sheet_rows(wb[sheet]):
            home_raw, away_raw = row.get("Home"), row.get("Away")
            if not home_raw or not away_raw:
                continue
            if finals:
                hg, ag = _f(row.get("HGFT")), _f(row.get("AGFT"))
            else:
                hg, ag = _f(row.get("HG")), _f(row.get("AG"))
            het, aet = _f(row.get("HGET")), _f(row.get("AGET"))
            hpen, apen = _f(row.get("HGP")), _f(row.get("HGP.1"))
            finished_raw = row.get("Finished")
            if hpen is not None or apen is not None:
                finish_type = "penalties"
            elif het is not None or aet is not None:
                finish_type = "extra_time"
            else:
                finish_type = "regulation"
            dirty = bool(finished_raw and str(finished_raw) in ("Penalties", "Extra time")
                         and finish_type == "regulation")
            h_id, a_id = mid(home_raw), mid(away_raw)
            oh, od, oa = odds_1x2(row, finals)
            odds_ok = all(v is not None and v > 1.0 for v in (oh, od, oa))
            is_host_home = bool(finals and host and canon_name(home_raw) == host)
            is_host_away = bool(finals and host and canon_name(away_raw) == host)
            neutral = False if not finals else (not is_host_home and not is_host_away)
            stage = "knockout" if (finals and finish_type != "regulation") else ("group_or_unknown" if finals else "qualifier")
            hxg, axg = _f(row.get("HxG")), _f(row.get("AxG"))
            hs, as_ = _f(row.get("HS")), _f(row.get("AS"))
            rec = {
                "source_sheet": sheet, "competition": row.get("Competition") or sheet,
                "season": season, "match_date": _date(row.get("Date")),
                "phase": "qualifier" if not finals else "finals", "stage": stage,
                "neutral_site": neutral, "is_host_home": is_host_home, "is_host_away": is_host_away,
                "home_name_raw": str(home_raw).strip(), "away_name_raw": str(away_raw).strip(),
                "home_team_id": h_id, "away_team_id": a_id,
                "home_goals_90": int(hg) if hg is not None else None,
                "away_goals_90": int(ag) if ag is not None else None,
                "finish_type": finish_type, "finished_label_raw": finished_raw,
                "dirty_finished_label": dirty,
                "home_goals_et": het, "away_goals_et": aet,
                "home_penalties": hpen, "away_penalties": apen,
                "odds_1x2_home": oh, "odds_1x2_draw": od, "odds_1x2_away": oa,
                "odds_1x2_available": odds_ok,
                "odds_source": ("avg_consensus" if (finals and _f(row.get("H-Avg")) is not None) or (not finals and _f(row.get("H_Avg")) is not None) else "max_or_book"),
                "ou_main_line": None, "ou_over_odds": None, "ou_under_odds": None, "ou_market_available": False,
                "ah_main_line": None, "ah_market_available": False,
                "home_xg": hxg, "away_xg": axg, "xg_available": hxg is not None and axg is not None,
                "home_shots": hs, "away_shots": as_,
                "home_sot": _f(row.get("HST")), "away_sot": _f(row.get("AST")),
                "home_corners": _f(row.get("HC")), "away_corners": _f(row.get("AC")),
                "home_fouls": _f(row.get("HF")), "away_fouls": _f(row.get("AF")),
                "stats_available": hs is not None,
                "pipeline_mode": "1X2_ONLY", "w1_full_pipeline_validated": False,
                "prematch_available": True,
                "result_available": hg is not None and ag is not None,
            }
            rows_out.append(rec)

    if unmapped:
        raise SystemExit(
            "FATAL: unmapped team names (no silent fallback). Add to config/w1_team_aliases.json:\n"
            + "\n".join(f"  {n} (x{c})" for n, c in sorted(unmapped.items()))
        )

    OUT_CSV.parent.mkdir(parents=True, exist_ok=True)
    with OUT_CSV.open("w", encoding="utf-8", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=FIELDS)
        w.writeheader()
        for r in rows_out:
            w.writerow(r)

    summary = build_summary(rows_out)
    OUT_SUMMARY.write_text(json.dumps(summary, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return summary


def build_summary(rows: list[dict[str, Any]]) -> dict[str, Any]:
    n = len(rows)
    by_sheet: dict[str, int] = {}
    for r in rows:
        by_sheet[r["source_sheet"]] = by_sheet.get(r["source_sheet"], 0) + 1
    q = [r for r in rows if r["phase"] == "qualifier"]
    return {
        "schema_version": "W1_INTERNATIONAL_COVERAGE_V1",
        "pipeline_mode": "1X2_ONLY",
        "w1_full_pipeline_validated": False,
        "total_rows": n,
        "by_sheet": by_sheet,
        "odds_1x2_available": sum(1 for r in rows if r["odds_1x2_available"]),
        "ou_market_available": sum(1 for r in rows if r["ou_market_available"]),
        "ah_market_available": sum(1 for r in rows if r["ah_market_available"]),
        "xg_available_total": sum(1 for r in rows if r["xg_available"]),
        "xg_available_qualifiers": sum(1 for r in q if r["xg_available"]),
        "qualifier_rows": len(q),
        "stats_available_total": sum(1 for r in rows if r["stats_available"]),
        "fouls_available_total": sum(1 for r in rows if r["home_fouls"] is not None),
        "dirty_finished_labels": [
            {"sheet": r["source_sheet"], "match": f"{r['home_name_raw']} vs {r['away_name_raw']}",
             "score": f"{r['home_goals_90']}-{r['away_goals_90']}", "finished_label_raw": r["finished_label_raw"]}
            for r in rows if r["dirty_finished_label"]
        ],
        "penalties_rows": sum(1 for r in rows if r["finish_type"] == "penalties"),
        "extra_time_rows": sum(1 for r in rows if r["finish_type"] == "extra_time"),
        "notes_cn": [
            "OU/AH 在该数据源缺失，pipeline_mode=1X2_ONLY，不能复现完整 W1 比分矩阵管线。",
            "90 分钟比分用于建模；加时/点球单列，仅作淘汰赛标记。",
            "finish_type 由 ET/点球字段推导，不信 Finished 标签；脏标签已定位。",
        ],
    }


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", type=Path, default=DEFAULT_XLSX)
    ap.add_argument("--rebuild-aliases", action="store_true")
    args = ap.parse_args()
    if not args.input.is_file():
        raise SystemExit(f"input workbook not found: {args.input}")
    if args.rebuild_aliases:
        payload = build_aliases(args.input)
        ALIASES.parent.mkdir(parents=True, exist_ok=True)
        ALIASES.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
        print(f"team aliases written: {len(payload['teams'])} teams -> {ALIASES.relative_to(ROOT)}")
        return 0
    summary = normalize(args.input)
    print(f"W1 international dataset normalized: rows={summary['total_rows']} sheets={summary['by_sheet']}")
    print(f"  odds_1x2={summary['odds_1x2_available']} xG(qual)={summary['xg_available_qualifiers']}/{summary['qualifier_rows']} "
          f"stats={summary['stats_available_total']} fouls={summary['fouls_available_total']} dirty_finished={len(summary['dirty_finished_labels'])}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
